#!/usr/bin/env python3
"""
生成企业标准格式的报销明细表（Excel）

使用方法：
    python create_expense_sheet.py expenses.json output.xlsx "公司名" "部门" "报销人"

expenses.json 格式：
[
    {"seq": 1, "date": "2026-03-11", "content": "打车", "amount": 62.89, "note": ""},
    ...
]
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import json
import sys

# 边框样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 列宽配置
COLUMN_WIDTHS = {'A': 12.25, 'B': 16.75, 'C': 39.5, 'D': 17.75, 'E': 20.33}


def create_expense_sheet(expenses, output_path, company, department,报销人):
    """
    生成报销明细表
    
    样式配置（基于企业标准模板）：
    - 标题行：宋体 18pt 加粗 居中
    - 表头行：宋体 12pt
    - 列标题行：宋体 12pt，有边框
    - 数据行：宋体 10pt，有边框
    - 合计行：宋体 12pt
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报销明细表"

    # 设置列宽
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # 行1: 标题
    ws['A1'] = "费用报销明细表"
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(name='宋体', size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 38.4

    # 行2: 归属公司
    ws['A2'] = "归属公司："
    ws['A2'].font = Font(name='宋体', size=12)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['B2'] = company
    ws.merge_cells('B2:E2')
    ws['B2'].font = Font(name='宋体', size=12)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 24.0

    # 行3: 归属部门 + 报销人
    ws['A3'] = f" 归属部门：{department}"
    ws.merge_cells('A3:D3')
    ws['A3'].font = Font(name='宋体', size=12)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws['E3'] = f"报销人:{报销人}"
    ws['E3'].font = Font(name='宋体', size=12)
    ws['E3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 24.9

    # 行4: 列标题
    headers = ["序号", "日期", "报销内容", "金额", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = h
        cell.font = Font(name='宋体', size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws.row_dimensions[4].height = 24.9

    # 数据行
    for idx, exp in enumerate(expenses):
        row = 5 + idx
        ws.row_dimensions[row].height = 24.9

        # 序号
        cell = ws.cell(row=row, column=1)
        cell.value = exp.get('seq', idx + 1)
        cell.font = Font(name='宋体', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 日期
        cell = ws.cell(row=row, column=2)
        cell.value = datetime.strptime(exp['date'], "%Y-%m-%d")
        cell.number_format = 'M/D/YYYY'
        cell.font = Font(name='宋体', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 报销内容
        cell = ws.cell(row=row, column=3)
        cell.value = exp['content']
        cell.font = Font(name='宋体', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 金额
        cell = ws.cell(row=row, column=4)
        cell.value = exp['amount']
        cell.font = Font(name='宋体', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

        # 备注
        cell = ws.cell(row=row, column=5)
        cell.value = exp.get('note', '')
        cell.font = Font(name='宋体', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 合计行
    sum_row = 5 + len(expenses)
    ws.cell(row=sum_row, column=1, value="合计：")
    ws.merge_cells(f'A{sum_row}:C{sum_row}')
    ws.cell(row=sum_row, column=1).font = Font(name='宋体', size=12)
    ws.cell(row=sum_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    ws.cell(row=sum_row, column=4, value=f"=SUM(D5:D{sum_row-1})")
    ws.cell(row=sum_row, column=4).font = Font(name='宋体', size=12)
    ws.cell(row=sum_row, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=sum_row, column=4).border = thin_border
    ws.row_dimensions[sum_row].height = 24.9

    wb.save(output_path)
    return sum(e['amount'] for e in expenses)


if __name__ == "__main__":
    if len(sys.argv) < 6:
        print("用法: python create_expense_sheet.py expenses.json output.xlsx 公司 部门 报销人")
        sys.exit(1)
    
    expenses_file = sys.argv[1]
    output_path = sys.argv[2]
    company = sys.argv[3]
    department = sys.argv[4]
    报销人 = sys.argv[5]
    
    with open(expenses_file, 'r', encoding='utf-8') as f:
        expenses = json.load(f)
    
    total = create_expense_sheet(expenses, output_path, company, department, 报销人)
    print(f"已生成: {output_path}")
    print(f"共 {len(expenses)} 项，合计: {total:.2f} 元")
