#!/usr/bin/env python3
"""Run the local product-video -> Xiaohongshu publish -> trace -> ledger pipeline."""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent
PRODUCT_VIDEO_SCRIPT = Path("/Users/luxiaochuan/fuxingdaoOPC/skills/product-promo-video/scripts/generate_campaign.py")
TRACE_SKILL_PATH = Path("/Users/luxiaochuan/fuxingdaoOPC/skills/xhs-publish-trace/SKILL.md")

if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

from mcp_http_client import MCPHttpClient  # noqa: E402


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Batch process 待完成 tasks into published Xiaohongshu videos.")
    parser.add_argument("--todo-dir", default="/Users/luxiaochuan/fuxingdaoOPC/待完成", help="Root folder containing task subfolders.")
    parser.add_argument("--output-root", default="/Users/luxiaochuan/fuxingdaoOPC/已落盘", help="Root folder for dated outputs.")
    parser.add_argument("--date", default="", help="Run date in YYYY-MM-DD. Defaults to local today.")
    parser.add_argument("--ratio", default="9:16", choices=["9:16", "16:9"], help="Video aspect ratio for product-promo-video.")
    parser.add_argument("--resolution", default="720p", choices=["480p", "720p", "1080p"], help="Video resolution.")
    parser.add_argument("--segment-duration", type=int, default=7, help="Duration per generated segment.")
    parser.add_argument("--platform", default="小红书", help="Target platform label passed into product-promo-video.")
    parser.add_argument("--visibility", default="公开可见", choices=["公开可见", "仅自己可见", "仅互关好友可见"], help="XHS visibility.")
    parser.add_argument("--trace-retries", type=int, default=3, help="How many keyword passes to use while tracing a note.")
    parser.add_argument("--dry-run", action="store_true", help="Generate videos and ledger, but skip publish and trace.")
    parser.add_argument("--force-video", action="store_true", help="Regenerate videos even when campaign-report.json already exists.")
    parser.add_argument("--force-publish", action="store_true", help="Republish even when pipeline-record.json already has publish data.")
    parser.add_argument("--force-trace", action="store_true", help="Retrace even when a note URL already exists.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    run_date = args.date or datetime.now().astimezone().date().isoformat()
    todo_dir = Path(args.todo_dir).expanduser().resolve()
    output_root = Path(args.output_root).expanduser().resolve()
    daily_dir = output_root / run_date
    daily_dir.mkdir(parents=True, exist_ok=True)

    tasks = discover_tasks(todo_dir)
    client = None if args.dry_run else MCPHttpClient()
    if client is not None:
        ensure_login(client)

    rows: list[dict[str, Any]] = []
    summary = {
        "run_date": run_date,
        "todo_dir": str(todo_dir),
        "output_root": str(output_root),
        "daily_dir": str(daily_dir),
        "trace_skill": str(TRACE_SKILL_PATH),
        "task_count": len(tasks),
        "dry_run": args.dry_run,
        "visibility": args.visibility,
        "tasks": [],
    }

    for task_dir in tasks:
        record = process_task(task_dir, daily_dir, args, client, run_date)
        rows.append(record)
        summary["tasks"].append(record)

    ledger_path = daily_dir / "daily-ledger.xlsx"
    write_workbook(ledger_path, rows, summary)
    (daily_dir / "pipeline-summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(json.dumps({
        "daily_dir": str(daily_dir),
        "ledger": str(ledger_path),
        "task_count": len(tasks),
        "dry_run": args.dry_run,
    }, ensure_ascii=False, indent=2))


def discover_tasks(todo_dir: Path) -> list[Path]:
    if not todo_dir.exists():
        raise SystemExit(f"待处理目录不存在: {todo_dir}")
    tasks = [path for path in sorted(todo_dir.iterdir()) if path.is_dir() and not path.name.startswith(".")]
    return tasks


def ensure_login(client: MCPHttpClient) -> None:
    response = client.call_tool("check_login_status", {}, timeout=30)
    text = response.content_text
    if "已登录" not in text and "Logged in" not in text:
        raise SystemExit(f"XHS MCP 未登录，无法继续:\n{text}")


def process_task(
    task_dir: Path,
    daily_dir: Path,
    args: argparse.Namespace,
    client: MCPHttpClient | None,
    run_date: str,
) -> dict[str, Any]:
    task_id = task_dir.name
    output_dir = daily_dir / task_id
    output_dir.mkdir(parents=True, exist_ok=True)
    record_path = output_dir / "pipeline-record.json"
    existing = load_json(record_path)

    row = base_row(task_dir, output_dir, run_date)
    if existing:
        row.update(existing)

    try:
        report = ensure_video(task_dir, output_dir, args)
        row.update(build_row_from_report(report, output_dir))

        payload = build_publish_payload(report, row["task_id"], args.visibility)
        row["publish_title"] = payload["title"]
        row["publish_body"] = payload["content"]
        row["publish_tags"] = ",".join(payload["tags"])
        row["visibility"] = payload["visibility"]

        if args.dry_run:
            row["publish_status"] = "skipped(dry-run)"
            row["trace_status"] = "skipped(dry-run)"
            row["status"] = "ready"
        else:
            if args.force_publish or not row.get("publish_result_text"):
                publish_result = client.call_tool("publish_with_video", payload, timeout=1800)
                row["publish_result_text"] = publish_result.content_text
                row["publish_status"] = "published"
            else:
                row["publish_status"] = row.get("publish_status") or "reused"

            if args.force_trace or not row.get("note_url"):
                trace = trace_note(client, payload, report, args.trace_retries)
                row.update(trace)
            else:
                row["trace_status"] = row.get("trace_status") or "reused"

            row["status"] = "completed" if row.get("note_url") else "published-awaiting-trace"
    except Exception as exc:  # noqa: BLE001
        row["status"] = "failed"
        row["error"] = str(exc)

    record_path.write_text(json.dumps(row, ensure_ascii=False, indent=2), encoding="utf-8")
    return row


def ensure_video(
    task_dir: Path,
    output_dir: Path,
    args: argparse.Namespace,
) -> dict[str, Any]:
    report_path = output_dir / "campaign-report.json"
    if report_path.exists() and not args.force_video:
        report = load_json(report_path)
        final_video = Path(report.get("final_video", ""))
        if report and final_video.exists() and campaign_is_complete(report):
            return report

    cmd = [
        sys.executable,
        str(PRODUCT_VIDEO_SCRIPT),
        "--input-dir",
        str(task_dir),
        "--platform",
        args.platform,
        "--ratio",
        args.ratio,
        "--resolution",
        args.resolution,
        "--segment-duration",
        str(args.segment_duration),
        "--out-dir",
        str(output_dir),
    ]
    run_command(cmd)
    report = load_json(report_path)
    if not report:
        raise RuntimeError(f"视频生成后缺少 campaign-report.json: {report_path}")
    if not campaign_is_complete(report):
        raise RuntimeError("campaign-report.json 缺少完整的两段视频或关键帧信息，拒绝降级为单段视频。")
    final_video = Path(report.get("final_video", ""))
    if not final_video.exists():
        raise RuntimeError(f"视频生成后缺少最终视频文件: {final_video}")
    return report


def campaign_is_complete(report: dict[str, Any]) -> bool:
    segment_1_url = clean_text(report.get("segment_1_url", ""))
    segment_2_url = clean_text(report.get("segment_2_url", ""))
    final_video = Path(report.get("final_video", ""))
    keyframes = report.get("keyframes", {}) or {}
    required_keyframes = ["seg1_start", "seg1_end", "seg2_start", "seg2_end"]
    if not segment_1_url or not segment_2_url:
        return False
    if not final_video.exists():
        return False
    for key in required_keyframes:
        path = Path(keyframes.get(key, ""))
        if not path.exists():
            return False
    return True


def build_publish_payload(
    report: dict[str, Any],
    task_id: str,
    visibility: str,
) -> dict[str, Any]:
    analysis = report.get("input_analysis", {}) or {}
    product_name = clean_text(analysis.get("product_name") or "")
    summary = clean_text(analysis.get("summary") or report.get("description") or "")
    selling_points = analysis.get("selling_points") or []
    spoken_1 = clean_text(report.get("segment_1_copy") or "")
    spoken_2 = clean_text(report.get("segment_2_copy") or "")
    keywords = normalize_tags(analysis.get("visual_keywords") or [], product_name, report.get("category") or "")

    title = shorten_title(product_name or summary or f"任务{task_id}")
    content_parts = [summary]
    for point in selling_points[:2]:
        if point:
            content_parts.append(f"• {clean_text(str(point))}")
    if spoken_1:
        content_parts.append(spoken_1)
    if spoken_2:
        content_parts.append(spoken_2)
    content = "\n".join(part for part in content_parts if part).strip()
    content = content[:900]

    return {
        "title": title,
        "content": content,
        "tags": keywords[:8],
        "video": str(Path(report["final_video"]).resolve()),
        "visibility": visibility,
    }


def trace_note(
    client: MCPHttpClient,
    payload: dict[str, Any],
    report: dict[str, Any],
    trace_retries: int,
) -> dict[str, Any]:
    queries = build_trace_queries(payload, report)[: max(trace_retries, 1)]
    for query in queries:
        search_data = client.call_tool_json(
            "search_feeds",
            {
                "keyword": query,
                "filters": {
                    "sort_by": "最新",
                    "note_type": "视频",
                },
            },
            timeout=120,
        )
        feeds = search_data.get("feeds", []) or []
        for feed in feeds:
            note_card = feed.get("noteCard", {}) or {}
            if note_card.get("type") != "video":
                continue
            if not matches_candidate(note_card, payload, report):
                continue
            feed_id = feed.get("id", "")
            xsec_token = feed.get("xsecToken", "")
            if not feed_id or not xsec_token:
                continue
            detail = client.call_tool_json(
                "get_feed_detail",
                {"feed_id": feed_id, "xsec_token": xsec_token},
                timeout=120,
            )
            note = (detail.get("data", {}) or {}).get("note", {}) or detail.get("note", {}) or {}
            interact = note.get("interactInfo", {}) or {}
            note_title = clean_text(note.get("title") or note_card.get("displayTitle") or "")
            if note_title and payload["title"] and payload["title"] not in note_title and note_title not in payload["title"]:
                if not matches_candidate(note_card, payload, report):
                    continue
            return {
                "trace_status": "found",
                "video_url": build_note_url(feed_id, xsec_token),
                "note_url": build_note_url(feed_id, xsec_token),
                "feed_id": feed_id,
                "xsec_token": xsec_token,
                "likes": interact.get("likedCount", ""),
                "collects": interact.get("collectedCount", ""),
                "comments": interact.get("commentCount", ""),
                "shares": interact.get("sharedCount", ""),
            }
    return {
        "trace_status": "not_found",
        "video_url": "",
        "note_url": "",
        "feed_id": "",
        "xsec_token": "",
        "likes": "",
        "collects": "",
        "comments": "",
        "shares": "",
    }


def build_trace_queries(payload: dict[str, Any], report: dict[str, Any]) -> list[str]:
    analysis = report.get("input_analysis", {}) or {}
    candidates = [
        payload.get("title", ""),
        clean_text(analysis.get("product_name") or ""),
        first_distinctive_phrase(payload.get("content", "")),
    ]
    deduped: list[str] = []
    for item in candidates:
        cleaned = clean_text(item)
        if cleaned and cleaned not in deduped:
            deduped.append(cleaned)
    return deduped


def matches_candidate(note_card: dict[str, Any], payload: dict[str, Any], report: dict[str, Any]) -> bool:
    analysis = report.get("input_analysis", {}) or {}
    haystack = " ".join(
        clean_text(part)
        for part in [
            note_card.get("displayTitle", ""),
            analysis.get("product_name", ""),
            report.get("description", ""),
        ]
        if part
    )
    needles = [
        clean_text(payload.get("title", "")),
        clean_text(analysis.get("product_name", "")),
    ]
    for needle in needles:
        if needle and needle in haystack:
            return True
    display_title = clean_text(note_card.get("displayTitle", ""))
    return not display_title


def base_row(task_dir: Path, output_dir: Path, run_date: str) -> dict[str, Any]:
    image_path = first_image(task_dir)
    copy_path = task_dir / "文案.md"
    return {
        "run_date": run_date,
        "task_id": task_dir.name,
        "product_name": "",
        "video_name": "",
        "video_url": "",
        "source_dir": str(task_dir),
        "image_path": str(image_path) if image_path else "",
        "copy_path": str(copy_path) if copy_path.exists() else "",
        "output_dir": str(output_dir),
        "report_path": "",
        "final_video": "",
        "category": "",
        "summary": "",
        "publish_title": "",
        "publish_body": "",
        "publish_tags": "",
        "visibility": "",
        "video_status": "",
        "publish_status": "",
        "trace_status": "",
        "note_url": "",
        "feed_id": "",
        "xsec_token": "",
        "likes": "",
        "collects": "",
        "comments": "",
        "shares": "",
        "publish_result_text": "",
        "status": "pending",
        "error": "",
    }


def build_row_from_report(report: dict[str, Any], output_dir: Path) -> dict[str, Any]:
    analysis = report.get("input_analysis", {}) or {}
    return {
        "report_path": str((output_dir / "campaign-report.json").resolve()),
        "final_video": str(Path(report.get("final_video", "")).resolve()),
        "product_name": clean_text(analysis.get("product_name") or ""),
        "video_name": clean_text(analysis.get("product_name") or ""),
        "category": clean_text(report.get("category") or ""),
        "summary": clean_text(analysis.get("summary") or report.get("description") or ""),
        "video_status": "generated",
    }


def write_workbook(path: Path, rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_tasks = wb.create_sheet("Tasks")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    summary_rows = [
        ("run_date", summary["run_date"]),
        ("todo_dir", summary["todo_dir"]),
        ("output_root", summary["output_root"]),
        ("daily_dir", summary["daily_dir"]),
        ("trace_skill", summary["trace_skill"]),
        ("task_count", summary["task_count"]),
        ("dry_run", str(summary["dry_run"])),
        ("visibility", summary["visibility"]),
    ]
    for index, (key, value) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=index, column=1, value=key)
        ws_summary.cell(row=index, column=2, value=value)
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 110

    columns = [
        "run_date",
        "task_id",
        "product_name",
        "video_name",
        "video_url",
        "source_dir",
        "image_path",
        "copy_path",
        "output_dir",
        "report_path",
        "final_video",
        "category",
        "summary",
        "publish_title",
        "publish_body",
        "publish_tags",
        "visibility",
        "video_status",
        "publish_status",
        "trace_status",
        "note_url",
        "feed_id",
        "xsec_token",
        "likes",
        "collects",
        "comments",
        "shares",
        "status",
        "error",
    ]
    for col_index, name in enumerate(columns, start=1):
        cell = ws_tasks.cell(row=1, column=col_index, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row_index, row in enumerate(rows, start=2):
        for col_index, name in enumerate(columns, start=1):
            cell = ws_tasks.cell(row=row_index, column=col_index, value=row.get(name, ""))
            cell.alignment = wrap_alignment

    ws_tasks.freeze_panes = "A2"
    ws_tasks.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(len(rows) + 1, 2)}"

    for col_index, name in enumerate(columns, start=1):
        width = max(len(name) + 2, 14)
        if name in {"source_dir", "image_path", "copy_path", "output_dir", "report_path", "final_video", "note_url", "video_url", "error"}:
            width = 28
        if name in {"summary", "publish_body", "xsec_token"}:
            width = 36
        ws_tasks.column_dimensions[get_column_letter(col_index)].width = width

    wb.save(path)


def run_command(cmd: list[str]) -> subprocess.CompletedProcess[str]:
    try:
        return subprocess.run(cmd, check=True, text=True, capture_output=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(
            f"命令失败: {' '.join(cmd)}\nSTDOUT:\n{exc.stdout}\nSTDERR:\n{exc.stderr}"
        ) from exc


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def first_image(task_dir: Path) -> Path | None:
    for path in sorted(task_dir.iterdir()):
        if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
            return path
    return None


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def shorten_title(value: str) -> str:
    title = clean_text(value).replace("#", "")
    if len(title) <= 20:
        return title
    return title[:20]


def normalize_tags(raw_keywords: Any, product_name: str, category: str) -> list[str]:
    values: list[str] = []
    if isinstance(raw_keywords, list):
        values.extend(str(item) for item in raw_keywords)
    elif raw_keywords:
        values.extend(re.split(r"[、,，\s]+", str(raw_keywords)))
    values.extend([product_name, category, "好物分享", "小红书视频"])
    deduped: list[str] = []
    for value in values:
        cleaned = clean_text(value).replace("#", "")
        if cleaned and cleaned not in deduped:
            deduped.append(cleaned)
    return deduped


def first_distinctive_phrase(content: str) -> str:
    text = clean_text(content)
    if "•" in text:
        first = text.split("•", 1)[0].strip()
        if first:
            return first[:20]
    return text[:20]


def build_note_url(feed_id: str, xsec_token: str) -> str:
    encoded = urllib.parse.quote(xsec_token, safe="")
    return f"https://www.xiaohongshu.com/explore/{feed_id}?xsec_token={encoded}&xsec_source=pc_feed"


if __name__ == "__main__":
    main()
