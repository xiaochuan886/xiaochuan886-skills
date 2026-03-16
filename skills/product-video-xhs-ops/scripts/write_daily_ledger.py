#!/usr/bin/env python3
"""Write/update the daily Excel ledger from one or more pipeline records."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLUMNS = [
    "run_date",
    "task_id",
    "product_name",
    "video_name",
    "video_url",
    "source_dir",
    "image_path",
    "copy_path",
    "output_dir",
    "report_path",
    "final_video",
    "category",
    "summary",
    "publish_title",
    "publish_body",
    "publish_tags",
    "visibility",
    "video_status",
    "publish_status",
    "trace_status",
    "note_url",
    "feed_id",
    "xsec_token",
    "likes",
    "collects",
    "comments",
    "shares",
    "status",
    "error",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Write daily Excel ledger from pipeline-record.json files.")
    parser.add_argument("--daily-dir", required=True, help="Absolute path to 已落盘/YYYY-MM-DD")
    parser.add_argument("--record", action="append", default=[], help="Absolute path to a pipeline-record.json file")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    daily_dir = Path(args.daily_dir).expanduser().resolve()
    if not daily_dir.exists():
        raise SystemExit(f"daily_dir 不存在: {daily_dir}")

    record_paths = [Path(item).expanduser().resolve() for item in args.record]
    if not record_paths:
        record_paths = sorted(daily_dir.glob("*/pipeline-record.json"))
    if not record_paths:
        raise SystemExit("未找到任何 pipeline-record.json")

    rows = [json.loads(path.read_text(encoding="utf-8")) for path in record_paths]
    summary = {
        "run_date": rows[0].get("run_date", daily_dir.name),
        "daily_dir": str(daily_dir),
        "task_count": len(rows),
        "completed": sum(1 for row in rows if row.get("status") == "completed"),
        "published_awaiting_trace": sum(1 for row in rows if row.get("status") == "published-awaiting-trace"),
        "failed": sum(1 for row in rows if row.get("status") == "failed"),
    }

    workbook = Workbook()
    ws_summary = workbook.active
    ws_summary.title = "Summary"
    ws_tasks = workbook.create_sheet("Tasks")

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    summary_rows = list(summary.items())
    for row_index, (key, value) in enumerate(summary_rows, start=1):
        ws_summary.cell(row=row_index, column=1, value=key)
        ws_summary.cell(row=row_index, column=2, value=value)
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 100

    for col_index, name in enumerate(COLUMNS, start=1):
        cell = ws_tasks.cell(row=1, column=col_index, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap

    for row_index, row in enumerate(rows, start=2):
        for col_index, name in enumerate(COLUMNS, start=1):
            cell = ws_tasks.cell(row=row_index, column=col_index, value=row.get(name, ""))
            cell.alignment = wrap

    ws_tasks.freeze_panes = "A2"

    for col_index, name in enumerate(COLUMNS, start=1):
        width = 16
        if name in {"source_dir", "image_path", "copy_path", "output_dir", "report_path", "final_video", "note_url", "video_url", "error"}:
            width = 28
        if name in {"summary", "publish_body", "xsec_token"}:
            width = 36
        ws_tasks.column_dimensions[get_column_letter(col_index)].width = width

    ledger_path = daily_dir / "daily-ledger.xlsx"
    summary_path = daily_dir / "pipeline-summary.json"
    workbook.save(ledger_path)
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({
        "ledger": str(ledger_path),
        "summary": str(summary_path),
        "records": [str(path) for path in record_paths],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
